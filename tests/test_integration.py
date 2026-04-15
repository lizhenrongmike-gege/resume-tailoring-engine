#!/usr/bin/env python3
"""Integration test: parse → simulate summaries → merge."""
import json
import os
import subprocess
import sys

import yaml

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..")
READ_SCRIPT = os.path.join(PROJECT_ROOT, "scripts", "batch_read_excel.py")
MERGE_SCRIPT = os.path.join(PROJECT_ROOT, "scripts", "batch_merge_summaries.py")
FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "sample_batch.xlsx")
SUMMARIES_DIR = os.path.join(PROJECT_ROOT, "output", "summaries")


def test_full_pipeline():
    """Parse Excel → simulate summary YAMLs → merge back to Excel."""

    # Step 1: Parse
    result = subprocess.run(
        [sys.executable, READ_SCRIPT, FIXTURE],
        capture_output=True, text=True,
        cwd=PROJECT_ROOT,
    )
    assert result.returncode == 0, f"Read script failed: {result.stderr}"
    entries = json.loads(result.stdout)
    assert len(entries) == 3

    # Step 2: Simulate sub-agent output (write summary YAMLs)
    for entry in entries:
        summary = {
            "company": entry["company"],
            "role_title": "Test Role",
            "slug": entry["slug"],
            "fit_score": "85%",
            "keywords_placed": 8,
            "keywords_total": 10,
            "keywords_missing": ["SomeSkill", "AnotherSkill"],
            "output_docx": f"output/doc/{entry['company']}_Test_Role_Resume.docx",
            "output_pdf": f"output/pdf/{entry['company']}_Test_Role_Resume.pdf",
            "pages": 1,
            "lines": 55,
            "brief_summary": f"Simulated summary for {entry['company']}. Good match.",
            "status": "SUCCESS",
        }
        with open(os.path.join(SUMMARIES_DIR, f"{entry['slug']}.yaml"), "w") as f:
            yaml.dump(summary, f)

    # Step 3: Merge
    result = subprocess.run(
        [sys.executable, MERGE_SCRIPT, FIXTURE],
        capture_output=True, text=True,
        cwd=PROJECT_ROOT,
    )
    assert result.returncode == 0, f"Merge script failed: {result.stderr}"

    # Step 4: Verify results Excel
    from openpyxl import load_workbook
    results_path = FIXTURE.replace(".xlsx", "_results.xlsx")
    assert os.path.exists(results_path), "Results file not created"

    wb = load_workbook(results_path)
    ws = wb.active

    for row_idx in [2, 3, 4]:
        cell = ws.cell(row=row_idx, column=3).value
        assert cell is not None, f"Row {row_idx} Column C is empty"
        assert "Simulated summary" in cell, f"Row {row_idx} unexpected: {cell}"

    print("PASS: Full pipeline integration test")

    # Cleanup
    os.remove(results_path)
    for entry in entries:
        yaml_path = os.path.join(SUMMARIES_DIR, f"{entry['slug']}.yaml")
        if os.path.exists(yaml_path):
            os.remove(yaml_path)
    manifest_path = os.path.join(SUMMARIES_DIR, "manifest.json")
    if os.path.exists(manifest_path):
        os.remove(manifest_path)


if __name__ == "__main__":
    try:
        test_full_pipeline()
        sys.exit(0)
    except (AssertionError, Exception) as e:
        print(f"FAIL: {e}")
        sys.exit(1)
