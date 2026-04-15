#!/usr/bin/env python3
"""Tests for batch_merge_summaries.py."""
import json
import os
import shutil
import subprocess
import sys
import tempfile

import yaml
from openpyxl import Workbook

SCRIPT = os.path.join(os.path.dirname(__file__), "..", "scripts", "batch_merge_summaries.py")
FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "sample_batch.xlsx")
SUMMARIES_DIR = os.path.join(os.path.dirname(__file__), "..", "output", "summaries")
PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..")


def setup():
    """Create manifest and sample summary YAMLs."""
    os.makedirs(SUMMARIES_DIR, exist_ok=True)

    manifest = {"2": "testcorp", "3": "testcorp_2", "4": "acme_inc"}
    with open(os.path.join(SUMMARIES_DIR, "manifest.json"), "w") as f:
        json.dump(manifest, f)

    success_summary = {
        "company": "TestCorp",
        "role_title": "Data Analyst",
        "slug": "testcorp",
        "fit_score": "80%",
        "keywords_placed": 10,
        "keywords_total": 12,
        "keywords_missing": ["BigQuery"],
        "output_docx": "output/doc/TestCorp_Data_Analyst_Resume.docx",
        "output_pdf": "output/pdf/TestCorp_Data_Analyst_Resume.pdf",
        "pages": 1,
        "lines": 55,
        "brief_summary": "Strong match. SQL and Python coverage excellent.",
        "status": "SUCCESS",
    }
    with open(os.path.join(SUMMARIES_DIR, "testcorp.yaml"), "w") as f:
        yaml.dump(success_summary, f)

    failed_summary = {
        "company": "TestCorp",
        "slug": "testcorp_2",
        "status": "FAILED",
        "error": "Sub-agent did not produce output",
    }
    with open(os.path.join(SUMMARIES_DIR, "testcorp_2.yaml"), "w") as f:
        yaml.dump(failed_summary, f)

    # acme_inc: no summary YAML (tests missing summary handling)


def run_script(xlsx_path):
    result = subprocess.run(
        [sys.executable, SCRIPT, xlsx_path],
        capture_output=True, text=True,
        cwd=PROJECT_ROOT,
    )
    return result


def test_results_file_created():
    """Merge should create a _results.xlsx file."""
    result = run_script(FIXTURE)
    assert result.returncode == 0, f"Script failed: {result.stderr}"
    results_path = FIXTURE.replace(".xlsx", "_results.xlsx")
    assert os.path.exists(results_path), "Results file not created"


def test_success_summary_written():
    """Column C of success row should contain brief_summary."""
    run_script(FIXTURE)
    from openpyxl import load_workbook
    results_path = FIXTURE.replace(".xlsx", "_results.xlsx")
    wb = load_workbook(results_path)
    ws = wb.active
    cell_c2 = ws.cell(row=2, column=3).value
    assert cell_c2 is not None, "Column C row 2 is empty"
    assert "Strong match" in cell_c2, f"Unexpected summary: {cell_c2}"


def test_failed_summary_written():
    """Column C of failed row should contain [FAILED] prefix."""
    run_script(FIXTURE)
    from openpyxl import load_workbook
    results_path = FIXTURE.replace(".xlsx", "_results.xlsx")
    wb = load_workbook(results_path)
    ws = wb.active
    cell_c3 = ws.cell(row=3, column=3).value
    assert cell_c3 is not None, "Column C row 3 is empty"
    assert "[FAILED]" in cell_c3, f"Expected [FAILED] prefix: {cell_c3}"


def test_missing_summary_handled():
    """Column C of missing summary row should contain [MISSING] prefix."""
    run_script(FIXTURE)
    from openpyxl import load_workbook
    results_path = FIXTURE.replace(".xlsx", "_results.xlsx")
    wb = load_workbook(results_path)
    ws = wb.active
    cell_c4 = ws.cell(row=4, column=3).value
    assert cell_c4 is not None, "Column C row 4 is empty"
    assert "[MISSING]" in cell_c4, f"Expected [MISSING] prefix: {cell_c4}"


def test_original_file_preserved():
    """Original Excel should not be modified."""
    from openpyxl import load_workbook
    wb = load_workbook(FIXTURE)
    ws = wb.active
    cell_c2 = ws.cell(row=2, column=3).value
    assert cell_c2 is None or cell_c2 == "", "Original file was modified"


def test_non_manifest_rows_are_cleared():
    """Rows excluded from the manifest should not retain stale Column C values."""
    os.makedirs(SUMMARIES_DIR, exist_ok=True)
    manifest_path = os.path.join(SUMMARIES_DIR, "manifest.json")
    summary_path = os.path.join(SUMMARIES_DIR, "testcorp.yaml")

    manifest_backup = None
    summary_backup = None
    if os.path.exists(manifest_path):
        with open(manifest_path) as f:
            manifest_backup = f.read()
    if os.path.exists(summary_path):
        with open(summary_path) as f:
            summary_backup = f.read()

    with open(manifest_path, "w") as f:
        json.dump({"20": "testcorp"}, f)

    with open(summary_path, "w") as f:
        yaml.dump(
            {
                "company": "TestCorp",
                "role_title": "Data Analyst",
                "slug": "testcorp",
                "fit_score": "80%",
                "keywords_placed": 10,
                "keywords_total": 12,
                "keywords_missing": ["BigQuery"],
                "output_docx": "output/doc/TestCorp_Data_Analyst_Resume.docx",
                "output_pdf": "output/pdf/TestCorp_Data_Analyst_Resume.pdf",
                "pages": 1,
                "lines": 55,
                "brief_summary": "Fresh summary for mapped row.",
                "status": "SUCCESS",
            },
            f,
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["company", "Job Description Text", "url"])
    for index in range(2, 20):
        worksheet.append([f"Placeholder {index}", f"Placeholder JD {index}", f"stale row {index} value"])
    worksheet.cell(row=20, column=1, value="TestCorp")
    worksheet.cell(row=20, column=2, value="JD text")
    worksheet.cell(row=20, column=3, value="stale row 20 value")
    worksheet.cell(row=21, column=1, value="nan")
    worksheet.cell(row=21, column=2, value="Skipped row should clear stale value")
    worksheet.cell(row=21, column=3, value="stale row 21 value")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        temp_path = handle.name

    try:
        workbook.save(temp_path)
        result = run_script(temp_path)
        assert result.returncode == 0, f"Script failed: {result.stderr}"

        from openpyxl import load_workbook
        results_path = temp_path.replace(".xlsx", "_results.xlsx")
        merged = load_workbook(results_path)
        merged_ws = merged.active

        assert merged_ws.cell(row=20, column=3).value == "Fresh summary for mapped row."
        assert merged_ws.cell(row=21, column=3).value in (None, ""), merged_ws.cell(row=21, column=3).value
    finally:
        results_path = temp_path.replace(".xlsx", "_results.xlsx")
        if os.path.exists(temp_path):
            os.remove(temp_path)
        if os.path.exists(results_path):
            os.remove(results_path)

        if manifest_backup is None:
            if os.path.exists(manifest_path):
                os.remove(manifest_path)
        else:
            with open(manifest_path, "w") as f:
                f.write(manifest_backup)

        if summary_backup is None:
            if os.path.exists(summary_path):
                os.remove(summary_path)
        else:
            with open(summary_path, "w") as f:
                f.write(summary_backup)


def cleanup():
    results_path = FIXTURE.replace(".xlsx", "_results.xlsx")
    if os.path.exists(results_path):
        os.remove(results_path)
    for fname in ["manifest.json", "testcorp.yaml", "testcorp_2.yaml"]:
        fpath = os.path.join(SUMMARIES_DIR, fname)
        if os.path.exists(fpath):
            os.remove(fpath)


def setup_module(module):
    setup()


def teardown_module(module):
    cleanup()


if __name__ == "__main__":
    setup()
    tests = [
        test_results_file_created,
        test_success_summary_written,
        test_failed_summary_written,
        test_missing_summary_handled,
        test_original_file_preserved,
        test_non_manifest_rows_are_cleared,
    ]
    passed = 0
    for test in tests:
        try:
            test()
            print(f"  PASS: {test.__name__}")
            passed += 1
        except AssertionError as e:
            print(f"  FAIL: {test.__name__}: {e}")
        except Exception as e:
            print(f"  ERROR: {test.__name__}: {e}")
    cleanup()
    print(f"\n{passed}/{len(tests)} tests passed")
    sys.exit(0 if passed == len(tests) else 1)
