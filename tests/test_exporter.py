#!/usr/bin/env python3
"""Tests for xlsx exporter."""
import os
import sys
import tempfile

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from jobscan.exporter import export_to_xlsx


def test_creates_xlsx_with_correct_headers():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    jobs = [
        {"company": "Stripe", "description": "Fraud role JD...",
         "url": "https://example.com/1", "location": "SF, CA"},
    ]
    export_to_xlsx(jobs, path)

    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    assert headers == ["company", "Job Description Text", "url", "location"]
    os.unlink(path)


def test_writes_job_data():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    jobs = [
        {"company": "Stripe", "description": "Investigate fraud...",
         "url": "https://stripe.com/jobs/1", "location": "San Francisco, CA"},
        {"company": "Adyen", "description": "KYC review...",
         "url": "https://adyen.com/jobs/2", "location": "SF, CA"},
    ]
    export_to_xlsx(jobs, path)

    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "Stripe"
    assert ws.cell(row=2, column=2).value == "Investigate fraud..."
    assert ws.cell(row=3, column=1).value == "Adyen"
    assert ws.max_row == 3
    os.unlink(path)


def test_empty_jobs_writes_header_only():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    export_to_xlsx([], path)

    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row == 1
    os.unlink(path)
