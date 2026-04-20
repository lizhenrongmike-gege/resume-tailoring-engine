import openpyxl
from app.services.excel_export import generate_batch_xlsx

def test_generates_valid_xlsx():
    jobs = [
        {"company": "OpenAI", "description": "Build ML systems", "url": "https://openai.com/apply", "location": "SF"},
        {"company": "Ramp", "description": "Build fintech", "url": "https://ramp.com/apply", "location": "NYC"},
    ]
    buffer = generate_batch_xlsx(jobs)
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == ["company", "Job Description Text", "url", "location"]
    assert ws.cell(2, 1).value == "OpenAI"
    assert ws.cell(3, 1).value == "Ramp"
    assert ws.max_row == 3  # header + 2 data rows

def test_empty_jobs_list():
    buffer = generate_batch_xlsx([])
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active
    assert ws.max_row == 1  # header only
