#!/usr/bin/env python3
"""Read a batch JD Excel file and output structured JSON.

Usage: python3 scripts/batch_read_excel.py <path_to_xlsx>

Input:  Excel with Column A = Company Name, Column B = JD text (header in row 1)
Output: JSON array to stdout: [{company, slug, row_index, jd_text}, ...]
Side effect: writes output/summaries/manifest.json mapping row_index → slug
"""
import json
import math
import os
import re
import sys

from openpyxl import load_workbook


def derive_slug(company_name):
    """Lowercase, replace non-alphanumeric with underscores, strip edges."""
    slug = company_name.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "_", slug)
    slug = slug.strip("_")
    return slug


def normalize_cell_text(value):
    """Return trimmed cell text, treating NaN placeholders as missing."""
    if value is None:
        return None

    if isinstance(value, float) and math.isnan(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    if text.lower() == "nan":
        return None

    return text


def main():
    if len(sys.argv) < 2:
        print("Usage: batch_read_excel.py <path_to_xlsx>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]

    if not os.path.exists(xlsx_path):
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    entries = []
    slug_counts = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        company = normalize_cell_text(row[0].value)
        jd_text = normalize_cell_text(row[1].value)

        if not company or not jd_text:
            continue

        row_index = row[0].row

        base_slug = derive_slug(company)
        if base_slug in slug_counts:
            slug_counts[base_slug] += 1
            slug = f"{base_slug}_{slug_counts[base_slug]}"
        else:
            slug_counts[base_slug] = 1
            slug = base_slug

        entries.append({
            "company": company,
            "slug": slug,
            "row_index": row_index,
            "jd_text": jd_text,
        })

    if not entries:
        print("Error: no data rows found in Excel", file=sys.stderr)
        sys.exit(1)

    # Write manifest
    os.makedirs("output/summaries", exist_ok=True)
    manifest = {str(e["row_index"]): e["slug"] for e in entries}
    with open("output/summaries/manifest.json", "w") as f:
        json.dump(manifest, f, indent=2)

    # Output JSON to stdout
    print(json.dumps(entries, indent=2))


if __name__ == "__main__":
    main()
