#!/usr/bin/env python3
"""Merge summary YAMLs back into the batch Excel file.

Usage: python3 scripts/batch_merge_summaries.py <path_to_xlsx>

Reads output/summaries/manifest.json for row→slug mapping.
Reads output/summaries/{slug}.yaml for each row's summary.
Writes results to {input_name}_results.xlsx (preserves original).
"""
import json
import os
import sys

import yaml
from openpyxl import load_workbook


def main():
    if len(sys.argv) < 2:
        print("Usage: batch_merge_summaries.py <path_to_xlsx>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    manifest_path = "output/summaries/manifest.json"

    if not os.path.exists(xlsx_path):
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(manifest_path):
        print(f"Error: manifest not found: {manifest_path}", file=sys.stderr)
        sys.exit(1)

    with open(manifest_path) as f:
        manifest = json.load(f)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).value = None

    for row_index_str, slug in manifest.items():
        row_idx = int(row_index_str)
        summary_path = f"output/summaries/{slug}.yaml"

        if not os.path.exists(summary_path):
            ws.cell(row=row_idx, column=3, value=f"[MISSING] No summary produced for {slug}")
            print(f"Warning: missing summary for {slug}", file=sys.stderr)
            continue

        with open(summary_path) as f:
            summary = yaml.safe_load(f)

        if summary.get("status") == "FAILED":
            error_msg = summary.get("error", "Unknown error")
            ws.cell(row=row_idx, column=3, value=f"[FAILED] {error_msg}")
        else:
            ws.cell(row=row_idx, column=3, value=summary.get("brief_summary", "No summary available"))

    # Write to _results.xlsx to preserve original
    base, ext = os.path.splitext(xlsx_path)
    output_path = f"{base}_results{ext}"
    wb.save(output_path)
    print(f"Results written to {output_path}")


if __name__ == "__main__":
    main()
